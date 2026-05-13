#!/usr/bin/env python3
"""Spec ingest helper — convert .docx / .xlsx → markdown for the /spec command.

Usage: spec_ingest.py <file1> [<file2> ...]

Output: prints a single markdown blob to stdout. Each source is delimited by a
`## Source: <path>` heading so the caller can attribute spans back.

Errors: prints per-file diagnostics to stderr; exits non-zero if any file
failed so the /spec command can show a fallback message.

Optional dependencies (graceful fallback if missing):
  - python-docx  (pip install python-docx)  — for .docx
  - openpyxl     (pip install openpyxl)     — for .xlsx
"""
import sys
from pathlib import Path


def _md_escape_cell(s: str) -> str:
    return (s or "").replace("|", "\\|").replace("\n", " ").strip()


def ingest_docx(path: Path):
    try:
        from docx import Document  # type: ignore
    except ImportError:
        return None, "python-docx not installed — try: pip install python-docx"

    try:
        doc = Document(str(path))
    except Exception as e:
        return None, f"failed to open: {e}"

    out = []
    W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"

    def cell_text(cell_el):
        # Direct-child paragraphs only — skip nested tables so outer-cell text
        # doesn't slurp inner-table content. Within a paragraph, iter(w:t) is
        # safe because runs/text spans are descendants of the paragraph itself.
        parts = []
        for child in cell_el:
            if child.tag == f"{W}p":
                parts.append("".join((t.text or "") for t in child.iter(f"{W}t")))
        return " ".join(p for p in parts if p.strip())

    for el in doc.element.body:
        tag = el.tag
        if tag == f"{W}p":
            text = "".join((t.text or "") for t in el.iter(f"{W}t"))
            if text.strip():
                out.append(text)
        elif tag == f"{W}tbl":
            # findall() returns direct children only — prevents inner-table rows
            # from being yielded as if they were rows of the outer table.
            rows = []
            for row in el.findall(f"{W}tr"):
                cells = [cell_text(c) for c in row.findall(f"{W}tc")]
                if cells:
                    rows.append(cells)
            if not rows:
                continue
            width = max(len(r) for r in rows)
            header = rows[0] + [""] * (width - len(rows[0]))
            out.append("")
            out.append("| " + " | ".join(_md_escape_cell(c) for c in header) + " |")
            out.append("| " + " | ".join("---" for _ in header) + " |")
            for r in rows[1:]:
                r = r + [""] * (width - len(r))
                out.append("| " + " | ".join(_md_escape_cell(c) for c in r) + " |")
            out.append("")

    return "\n\n".join(out).strip() + "\n", None


def ingest_xlsx(path: Path):
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return None, "openpyxl not installed — try: pip install openpyxl"

    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as e:
        return None, f"failed to open: {e}"

    MAX_ROWS = 5000  # guard against accidental "wrong file" / 500k-row sheets

    out = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out.append(f"### Sheet: {sheet_name}\n")
        rows = []
        truncated = False
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in row]
            while cells and not cells[-1].strip():
                cells.pop()
            if cells:
                rows.append(cells)
                if len(rows) > MAX_ROWS:
                    truncated = True
                    break
        if not rows:
            out.append("_(empty)_\n")
            continue

        width = max(len(r) for r in rows)
        # Header = first row with ≥2 non-empty cells (skips merged title rows
        # like "Pricing Sheet 2026" sitting alone above the real header).
        # Fall back to row 0 if every row has only one non-empty cell.
        header_idx = 0
        for i, r in enumerate(rows):
            if sum(1 for c in r if c.strip()) >= 2:
                header_idx = i
                break
        header = rows[header_idx] + [""] * (width - len(rows[header_idx]))
        out.append("| " + " | ".join(_md_escape_cell(c) for c in header) + " |")
        out.append("| " + " | ".join("---" for _ in header) + " |")
        for r in rows[header_idx + 1:]:
            r = r + [""] * (width - len(r))
            if not any(c.strip() for c in r):
                continue
            out.append("| " + " | ".join(_md_escape_cell(c) for c in r) + " |")
        if truncated:
            out.append(f"\n_(truncated at {MAX_ROWS} rows — re-export a smaller range if needed)_")
        out.append("")

    return "\n".join(out).rstrip() + "\n", None


def main(argv):
    if len(argv) < 2:
        print("Usage: spec_ingest.py <file1> [<file2> ...]", file=sys.stderr)
        return 2

    any_error = False
    chunks = []
    for arg in argv[1:]:
        path = Path(arg)
        if not path.exists():
            print(f"[error] {arg}: file not found", file=sys.stderr)
            any_error = True
            continue
        ext = path.suffix.lower()
        if ext == ".docx":
            content, err = ingest_docx(path)
        elif ext == ".xlsx":
            content, err = ingest_xlsx(path)
        else:
            print(f"[error] {arg}: unsupported extension '{ext}' (only .docx/.xlsx)", file=sys.stderr)
            any_error = True
            continue

        if err:
            print(f"[error] {arg}: {err}", file=sys.stderr)
            any_error = True
            continue

        chunks.append(f"\n---\n## Source: `{path}`\n\n{content}")

    if chunks:
        sys.stdout.write("".join(chunks))

    return 1 if any_error else 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
